import sqlite3
import PySimpleGUI as sg
from wakeonlan import send_magic_packet
from ipscan import Func
import openpyxl

def get_computers():
    wb = openpyxl.load_workbook(filename=r"\\FS04\Files_UD_Wesola$\UG\WIN\komputery\db.xlsx")
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == 'DB':
            break
    wb.active = s
    ws = wb.active
    for c in range(1, ws.max_column+1):
        if ws.cell(row=1, column=c).value == "MAC Address":
            break
    computers_offline = []
    for r in range(2, ws.max_row + 1):
        if not Func.detect_on(ws.cell(row=r, column=2).value):
            computers_offline.append({'name':ws.cell(row=r, column=3).value,
                              'user': ws.cell(row=r, column=5).value,
                              'ip': ws.cell(row=r, column=2).value,
                              'mac': ws.cell(row=r, column=4).value,
                              })
    '''
        WITH cte as (
            SELECT *, ROW_NUMBER() OVER (
                PARTITION BY name ORDER BY time desc) as rn
            from computers as c
            )
            select name, user, ip, mac, time from cte where rn=1
            order by name
    '''
    return computers_offline

if __name__ == '__main__':
    computers = get_computers()
    column = []
    for c in computers:
        column.append([sg.Text(c['name'],size=(14,1)), sg.Text(c['user'],size=(10,1)), sg.Text(c['ip'],size=(11,1)), sg.Text(c['mac'],size=(14,1)), sg.Button('Wake', key="BUTTON-" + c['mac'])])
    layout = [[sg.Column(column, scrollable=True, size=(530, 800))]]
    window = sg.Window("WakeOnLAN", layout)
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break
        if event[:7] == "BUTTON-":
            mac = event[7:]
            send_magic_packet(mac)
