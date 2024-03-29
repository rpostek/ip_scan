from typing import Callable
import subprocess
import csv
import os
import pathlib
import PySimpleGUI as sg
import socket
import ipaddress
import sys
import tempfile
from pefile import PE
import threading
import time
import json
import openpyxl
import datetime
import sqlite3
import yaml
import re


class Property:
    def __init__(self, name: str, func: Callable, description: str='', enabled: bool=True) -> None:
        self.name = name
        self.func = func
        self.description = description
        self.enabled = enabled

class Func:
    @staticmethod
    def get_processor_data(ip: str) -> dict:
        data = dict()
        try:
            r = Func.runPSjson(f'gwmi win32_Processor -comp {ip}')
            data['Processor'] = r['Name']
            return data
        except:
            pass
        return data

    @staticmethod
    def get_network_data(ip: str) -> dict:
        data = dict()
        try:
            #r = Func.runPSjson(f"gwmi -ClassName Win32_NetworkAdapterConfiguration -Filter \"IPEnabled='True'\" -comp {ip} |  Where-Object {{$_.PNPDeviceID -like 'PCI*'}} | select *")
            #r = Func.runPSjson(f"gwmi -ClassName Win32_NetworkAdapter -comp {ip} |  Where-Object {{($_.PNPDeviceID -like 'PCI*') -and ($_.NetEnabled )}} | select *")
            r = Func.runPSjson(f"Get-NetNeighbor {ip}")
            data['MAC Address'] = r['LinkLayerAddress']
            return data
        except:
            pass
        return data

    '''@staticmethod
    def get_WoL(ip: str) -> dict:
        data = dict()
        try:
            r = Func.runPSjson(f"$wol = gwmi -Class Win32_NetworkAdapter -Filter \"netenabled = 'true'\" -ComputerName {ip} |  Where-Object {{$_.PNPDeviceID -like 'PCI*'}}; gwmi MSPower_DeviceWakeEnable -Namespace root\wmi -ComputerName {ip}| where " + "{$_.instancename -match [regex]::escape($wol.PNPDeviceID) } | select -Property Enable")
            data['WakeOnLan'] = r['Enable']
            return data
        except:
            pass
        return data
    '''

    @staticmethod
    def get_bios_data(ip: str) -> dict:
        data = dict()
        try:
            r = Func.runPSjson(f'gwmi Win32_BIOS -comp {ip} | select *')
            data['BIOS'] = r['SMBIOSBIOSVersion']
            data['BIOS date'] = r['ReleaseDate'][:4] + '-' + r['ReleaseDate'][4:6] + '-' + r['ReleaseDate'][6:8]
            data['SerialNumber'] = r['SerialNumber']
            return data
        except:
            pass
        return data

    @staticmethod
    def get_computer_data(ip: str) -> dict:
        data = dict()
        try:
            r = Func.runPSjson(f'gwmi win32_computersystem -comp {ip}')
            data['User Name'] = r['UserName']
            data['Name'] = r['Name']
            data['System Family'] = r['SystemFamily']
            data['Logical Processors'] = r['NumberOfLogicalProcessors']
            data['Memory'] = f"{float(r['TotalPhysicalMemory']) / (2 ** 30):2.2f} GB"
            data['Model'] = r['Model']
            data['Manufacturer'] = r['Manufacturer']
            return data
        except:
            pass
        return data

    @staticmethod
    def get_disk_data(ip: str) -> dict:
        data = dict()
        try:
            r = Func.runPSjson(f'gwmi win32_logicaldisk -comp {ip} | where-object -Property DeviceID -eq "C:"')
            data['FreeSpace'] = f"{float(r['FreeSpace']) / (2 ** 30):2.2f} GB"
            data['DiskSize'] = f"{float(r['Size']) / (2 ** 30):2.2f} GB"
            return data
        except:
            pass
        return data



    @staticmethod
    def get_drivers_data(ip: str) -> dict:
        data = dict()
        try:
            r = Func.runPSjson(f'Get-WmiObject Win32_PnPSignedDriver -computer {ip} | Where-Object -Property DeviceName -EQ "Realtek PCIe GBE Family Controller"')
            data['Ethernet Driver'] = r['DriverVersion'] + ' ' + r['DeviceName']
            return data
        except:
            pass
        return data


    @staticmethod
    def get_monitor_data(ip: str) -> dict:
        data = dict()
        try:
            psqry = f'$dt = gwmi -Namespace root\\wmi -Class wmiMonitorID -comp {ip};'\
                    + r" $name = '';foreach ($i in $dt.UserFriendlyName) {$name += [char[]]$i};"\
                    + r" $sn = '';foreach ($i in $dt.SerialNumberID) {$sn += [char[]]$i};"\
                    + r" @{'Monitor YearOfManufacture'=$dt.YearOfManufacture;'Monitor Name'=$name;'Monitor SN'=$sn}"
            data = Func.runPSjson(psqry)
            data['Monitor SN'] = data['Monitor SN'].replace('\x00','')
            data['Monitor Name'] = data['Monitor Name'].replace('\x00','')
            return data
        except:
            pass
        return data

    @staticmethod
    def get_os_version(ip: str) -> dict:
        OS_VERSIONS = {
            '10.0.22000': 'Windows 11',
            '10.0.19045': 'Windows 10 (22H2)',
            '10.0.19044': 'Windows 10 (21H2)',
            '10.0.19043': 'Windows 10 (21H1)',
            '10.0.19042': 'Windows 10 (20H2)',
            '10.0.19041': 'Windows 10 (2004)',
            '10.0.18363': 'Windows 10 (1909)',
            '10.0.18362': 'Windows 10 (1903)',
            '10.0.17763': 'Windows 10 (1809)',
            '10.0.17134': 'Windows 10 (1803)',
            '10.0.16299': 'Windows 10 (1709)',
            '10.0.15063': 'Windows 10 (1703)',
            '10.0.14393': 'Windows 10 (1607)',
            '10.0.10586': 'Windows 10 (1511)',
            '10.0.10240': 'Windows 10',
            '6.3.9600': 'Windows 8.1 (Update 1)',
            '6.3.9200': 'Windows 8.1',
            '6.2.9200': 'Windows 8',
            '6.1.7601': 'Windows 7 SP1',
            '6.1.7600': 'Windows 7'
        }
        data = dict()
        try:
            r = Func.runPSjson(f'Get-WmiObject Win32_OperatingSystem -ComputerName {ip}')
            # for key in r.keys():
            #    print(key, ':', r[key])
            data['OS'] = r['Caption'] + ' ' + r['OSArchitecture']
            data['OS Version'] = OS_VERSIONS.get(r['Version'], 'build '+ r['Version'])
            # data['SerialNumber'] = r['SerialNumber']
            return data
        except:
            pass
        return data

    @staticmethod
    def get_office_version(ip: str) -> dict:
        data=dict()
        try:
            r = Func.runPSjson(f'Get-WmiObject win32_product -ComputerName {ip} |  ' + 'where{$_.Name -like "Microsoft Office Standard*" -or $_.Name -like "Microsoft Office Professional*"} | select Name,Version')
            # for key in r.keys():
            #    print(key, ':', r[key])
            data['Office'] = r['Name'] # + ' (' + r['Version'] +')'
            return data
        except:
            pass
        return data

    @staticmethod
    def get_software_installed(ip: str) -> dict:
        data=dict()
        try:
            r = Func.runPSjson(f'Get-WmiObject win32_product -ComputerName {ip} | ' + 'select Name,Version,Vendor')
            data['Software'] = ''
            for item in r:
                data['Software'] = data['Software'] + item['Name']  + ' (' + item['Version'] +')\n'
                item_found = re.search('CorelDRAW.*Draw',item['Name'])
                if item_found:
                    data['Corel'] = item['Name']  + ' (' + item['Version'] +')'
                item_found = re.search('Microsoft Office Standard.*',item['Name'])
                if item_found:
                    data['Office'] = item['Name']  + ' (' + item['Version'] +')'
                item_found = re.search('Microsoft Office Professional.*',item['Name'])
                if item_found:
                    data['Office'] = item['Name']  + ' (' + item['Version'] +')'
                item_found = re.search('ABBYY PDF Transformer.*',item['Name'])
                if item_found:
                    data['ABBYY PDF Transformer'] = item['Name']  + ' (' + item['Version'] +')'
                item_found = re.search('Java \\d.*',item['Name'])
                data_java = ''
                if item_found:
                    data_java = item['Vendor'] + ' ' + item['Name']  + ' (' + item['Version'] +')'
                    data['Java'] = data_java
                item_found = re.search('.*JRE \\d.*', item['Name'])
                if item_found:
                    data['Java'] = data_java + item['Vendor'] + ' ' + item['Name'] + ' (' + item['Version'] + ')'
            return data
        except:
            pass
        return data


    @staticmethod
    def runPSjson(cmd: str) -> str:
        try:
            ps = subprocess.Popen(["powershell", "-Command", cmd + ' | ConvertTo-json'],
                                  stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                                  creationflags=subprocess.CREATE_NO_WINDOW)
            data, errs = ps.communicate(timeout=120) #[0]
            data = data.decode(encoding='cp852')
            return json.loads(data)
        except subprocess.TimeoutExpired as e:
            ps.kill()
            data, errs = ps.communicate() #[0]
            print(errs)
            return dict()

    @staticmethod
    def get_time_source(ip: str) -> dict:
        data = dict()
        try:
            #r = Func.runPS(f'w32tm /query /status /computer:{ip} | Select-String -Pattern "^Source:.*"')
            #data['Time Source'] = r['Line'].replace('Source: ', '')
            r = Func.runPSjson(f'w32tm /query /status /computer:{ip}')
            data_temp=dict()
            for parm in r:
                try:
                    pair = parm.split(':', 1)
                    data_temp[pair[0]] = pair[1].strip()
                except:
                    pass
            data['Time Source'] = data_temp['Source'] + ' / ' + data_temp['Last Successful Sync Time']
            return data
        except:
            pass
        return data

    @staticmethod
    def get_last_user(ip: str) -> dict:
        try:
            last = (None, 0)
            for user in os.scandir(r'\\' + ip + r'\c$\Users'):
                if user.name == 'All Users':
                    continue
                moje_konto = pathlib.Path('\\\\' + ip + '\\c$\\Users\\' + user.name + '\\Desktop\\Moje konto.url')
                if moje_konto.exists():
                    if moje_konto.stat().st_mtime > last[1]:
                        last = user.name, moje_konto.stat().st_mtime
                    #logs.append((user.name, datetime.datetime.fromtimestamp(moje_konto.stat().st_mtime)))
                    #print(user, datetime.datetime.fromtimestamp(moje_konto.stat().st_mtime).isoformat())
            return {'Last Logged User': last[0]}
        except PermissionError:
            return {'Last Logged User': ''}


    @staticmethod
    def get_last_doc_time(ip: str) -> dict:
        try:
            last = (None, 0)
            for user in os.scandir(r'\\' + ip + r'\c$\Users'):
                if user.name == 'All Users':
                    continue
                recent_path = pathlib.Path('\\\\' + ip + '\\c$\\Users\\' + user.name + '\\AppData\\Roaming\\Microsoft\\Office\\Ostatnie')
                if recent_path.exists():
                    if recent_path.stat().st_mtime > last[1]:
                        last = user.name, recent_path.stat().st_mtime
            return {'Last Doc Time': datetime.date.fromtimestamp(last[1]).isoformat()}
        except PermissionError:
            return {'Last Doc Time': ''}



    @staticmethod
    def detect_on(ip: str) -> bool:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(0.2)
        try:
            s.connect((ip, 445))
            return True
        except socket.error as e:
            return False
        s.close()

    @staticmethod
    def get_version(path):
        '''podaje wersję pliku podanego w path'''
        pe = PE(path)
        if not 'VS_FIXEDFILEINFO' in pe.__dict__:
            return
        if not pe.VS_FIXEDFILEINFO:
            return
        verinfo = pe.VS_FIXEDFILEINFO[0]
        wersja_pliku = (verinfo.FileVersionMS >> 16, verinfo.FileVersionMS & 0xFFFF,
                             verinfo.FileVersionLS >> 16, verinfo.FileVersionLS & 0xFFFF)
        wersja_produktu = (verinfo.ProductVersionMS >> 16, verinfo.ProductVersionMS & 0xFFFF,
                                verinfo.ProductVersionLS >> 16, verinfo.ProductVersionLS & 0xFFFF)
        if wersja_produktu[0] == wersja_pliku[0] and wersja_produktu[1] == wersja_pliku[1] and wersja_produktu[2] == wersja_pliku[2]:
            wersja = ".".join(str(i) for i in wersja_pliku)
        else:
            wersja = str(wersja_produktu[0]) + '.' + ".".join(str(i) for i in wersja_pliku)
        return wersja

    @staticmethod
    def get_chrome_version(ip):
        wersja = ''
        try:
            if not isinstance(ip, str):
                ip=ip.__str__()
            filename = f'\\\\{ip}\\c$\\Program Files\\Google\\Chrome\\Application\\chrome.exe'
            if os.path.isfile(filename):
                wersja = Func.get_version(filename)
            filename = f'\\\\{ip}\\c$\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe'
            if os.path.isfile(filename):
                wersja = Func.get_version(filename)
        except:
            pass
        return {"Chrome Version": wersja}

    @staticmethod
    def get_firefox_version(ip):
        wersja = ''
        try:
            if not isinstance(ip, str):
                ip=ip.__str__()
            filename = f'\\\\{ip}\\c$\\Program Files\\Mozilla Firefox\\firefox.exe'
            if os.path.isfile(filename):
                wersja = Func.get_version(filename)
            wersja = Func.get_firefox_config(ip) + wersja
        except:
            pass
        return {"Firefox Version": wersja}


    @staticmethod
    def get_firefox_config(ip):
        cfg = '-'
        filename = f'\\\\{ip}\\c$\\Program Files\\Mozilla Firefox\\mozilla.cfg'
        if os.path.isfile(filename):
            if datetime.datetime.fromtimestamp(os.path.getmtime(filename)) > datetime.datetime(year=2022, month=5, day=1):
                cfg = '*'
        js = '-'
        filename = f'\\\\{ip}\\c$\\Program Files\\Mozilla Firefox\\defaults\\pref\\autoconfig.js'
        if os.path.isfile(filename):
            if datetime.datetime.fromtimestamp(os.path.getmtime(filename)) > datetime.datetime(year=2022, month=5, day=1):
                js = '*'
        return cfg + js


    @staticmethod
    def get_irfanview_version(ip):
        wersja = ''
        try:
            if not isinstance(ip, str):
                ip=ip.__str__()
            filename = f'\\\\{ip}\\c$\\Program Files (x86)\\IrfanView\\i_view32.exe'
            if os.path.isfile(filename):
                wersja = Func.get_version(filename)
        except:
            wersja = '---'
        return {"irfanview Version": wersja}

    @staticmethod
    def get_7zip_version(ip):
        wersja = ''
        try:
            if not isinstance(ip, str):
                ip=ip.__str__()
            filename = f'\\\\{ip}\\c$\\Program Files\\7-Zip\\7zFM.exe'
            if os.path.isfile(filename):
                wersja = Func.get_version(filename)
            filename = f'\\\\{ip}\\c$\\Program Files (x86)\\7-Zip\\7zFM.exe'
            if os.path.isfile(filename):
                wersja = Func.get_version(filename)
                wersja  = wersja + ' (32-bit)'
        except:
            wersja = '---'
        return {"7-Zip Version": wersja}

    @staticmethod
    def exists_printconfig(ip):
        if not isinstance(ip, str):
            ip=ip.__str__()
        if os.path.isfile(r'\\' + ip + r'\c$\Windows\System32\spool\drivers\x64\3\Printconfig.dll' ):
            return {'PrintConfig.dll': 'OK'}
        else:
            return {'PrintConfig.dll': '-'}

properties = (
    Property('Name', Func.get_computer_data, 'nazwa komputera', False),
    Property('MAC Address', Func.get_network_data, 'adres MAC aktywnej karty sieciowej', True),
    #Property('WakeOnLan', Func.get_WoL, 'czy włączona jest funkcja Wake on LAN', False),
    Property('Last Logged User', Func.get_last_user, 'ostatni zalogowany użytkownik', True),
    Property('User Name', Func.get_computer_data, 'zalogowany użytkownik', False),
    Property('OS', Func.get_os_version, 'system operacyjny', False),
    Property('OS Version', Func.get_os_version, 'wersja (wydanie) systemu operacyjnego', False),
    Property('Manufacturer', Func.get_computer_data, 'producent komputera', False),
    Property('System Family', Func.get_computer_data, 'model komputera', False),
    Property('Model', Func.get_computer_data, 'model komputera', False),
    Property('SerialNumber', Func.get_bios_data, 'numer seryjny komputera', False),
    Property('BIOS', Func.get_bios_data, 'wersja BIOSu', False),
    Property('BIOS date', Func.get_bios_data, 'data wydania BIOSu', False),
    Property('Processor', Func.get_processor_data, 'typ procesora', False),
    Property('Logical Processors', Func.get_computer_data, 'liczba procesorów logicznych', False),
    Property('Memory', Func.get_computer_data, 'całkowita pamięć fizyczna komputera', False),
    Property('Time Source', Func.get_time_source, 'nazwa serwera synchronizacji czasu i czas ostatniej synchronizacji', False),
    Property('Chrome Version', Func.get_chrome_version, 'wersja przeglądarki chrome', False),
    Property('Firefox Version', Func.get_firefox_version, 'wersja przeglądarki Firefox', False),
    Property('PrintConfig.dll', Func.exists_printconfig, 'obecnośc pliku PrintConfig.dll potrzebnego do drukowania z aplikacji UPW', False),
#    Property('Office', Func.get_office_version, 'wersja programu Office', False),
    Property('Office', Func.get_software_installed, 'wersja programu Office', False),
    Property('Last Doc Time', Func.get_last_doc_time, 'czas ostatniego otwartego dokmentu Office', False),
    Property('Corel', Func.get_software_installed, 'wersja programu Corel', False),
    Property('ABBYY PDF Transformer', Func.get_software_installed, 'wersja programu ABBYY PDF Transformer', False),
    Property('Java', Func.get_software_installed, 'wersja programu Java', False),
    Property('irfanview Version', Func.get_irfanview_version, 'wersja programu irfanview', False),
    Property('7-Zip Version', Func.get_7zip_version, 'wersja programu 7-Zip', False),
    Property('Software', Func.get_software_installed, 'zainstalowane oprogramowanie', False),
    Property('Monitor YearOfManufacture', Func.get_monitor_data, 'rok produkcji monitora', False),
    Property('Monitor Name', Func.get_monitor_data, 'model monitora', False),
    Property('Monitor SN', Func.get_monitor_data, 'numer seryjny monitora', False),
    Property('Ethernet Driver', Func.get_drivers_data, 'wesja sterownika Realtek PCIe GBE Family Controller', False),
    Property('DiskSize', Func.get_disk_data, 'pojemność dysku C:', False),
    Property('FreeSpace', Func.get_disk_data, 'wolne miejsce na dysku C:', False),
)

def get_params():
    global window
    my_ip = get_my_ip()
    sg.theme('LightBrown1')
    #layout = [[sg.Text("Parametry do wyświetlenia:")]]
    layout=[]
    columns = []
    column = []
    column_len = 8
    for pr_no, pr in enumerate(properties):
        if not pr_no % column_len:
            column = []
        column.append([sg.Checkbox(text=pr.name, tooltip=pr.description, default=pr.enabled, key='property-' + pr.name),])
        if len(column) == column_len:
            columns.append(sg.Column(column))
            if pr_no != len(properties) -1:
                columns.append(sg.VerticalSeparator(color='light grey'))
            column = []
    if len(column):
        columns.append(sg.Column(column, vertical_alignment='top'))
    layout.append(columns)

    layout.append([sg.Checkbox(text='export to Excel', default=False, key='save-xlsx', tooltip='Zapis wyników wyszukiwania do pliku Excela.')],)
    layout.append([sg.Submit("OK", pad=((20, 10), 3)),])
    layout.append([sg.HorizontalSeparator()], )
    layout.append([sg.Checkbox(text='use the following addr range', default=False, key='use-iprange',
                               tooltip='Użycie poiższego zakresu adresów zamiast podanego w pliku konfiguracujnym.')])
    layout.append([sg.Text(f'IP addr range ', pad=(1,3)),
                   sg.InputText(size=(3, 1), default_text=my_ip[0], key='ips1', pad=(1, 3)),
                   sg.InputText(size=(3, 1), default_text=my_ip[1], key='ips2', pad=(1, 3)),
                   sg.InputText(size=(3, 1), default_text=my_ip[2], key='ips3', pad=(1, 3)),
                   sg.InputText(size=(3, 1), default_text='1', key='ips4', pad=(1, 3)),
                   sg.Text(f' - ', pad=(1,3)),
                   sg.InputText(size=(3,1), default_text=my_ip[0], key='ipe1', pad=(1,3)),
                   sg.InputText(size=(3,1), default_text=my_ip[1], key='ipe2', pad=(1,3)),
                   sg.InputText(size=(3,1), default_text=my_ip[2], key='ipe3', pad=(1,3)),
                   sg.InputText(size=(3,1), default_text='170', key='ipe4', pad=(1,3))])
    #layout.append([sg.HorizontalSeparator()],)
    layout.append([sg.ProgressBar(1, orientation='h', size=(20, 20), key='progress'),],)
    #layout.append([sg.HorizontalSeparator()], )
    layout.append([ sg.Button("Update database", button_color='red3', tooltip='Odpytanie komputerów w zakresie wszystkich właściwosći i aktualizacja bazy danych.\nTrwa długo.\n Można zawęzić zakres adresacji opcją powyżej.'),])
    #layout.append([sg.ProgressBar(1, orientation='h', size=(20, 20), key='progress')],)
    window = sg.Window("ipscanner 2.0 beta", layout)
    while True:
        event, values = window.read()
        if event == "OK":
            # TODO disable controls of the form
            window.Element('OK').update(disabled=True)
            break
        if event == sg.WIN_CLOSED:
            sys.exit(0)
            # TODO close running threads
        if event == "Update database":
            window.Element("Update database").update(disabled=True)
            return {'Update database': True}

    #window.close()
    prop_list = [k.replace('property-', '') for k,v in values.items() if v and k.startswith('property-')]
    fun_list = {p.func for p in properties if p.name in prop_list}  #lista funkcji do wykonania
    prop_list = ['No', 'IP'] + prop_list  + ['Time',] #lista kolumn do wyświetlenia
    return {'property_list': prop_list, 'function_list': fun_list, 'use-iprange': values['use-iprange'],
            'ips1': values['ips1'], 'ips2': values['ips2'], 'ips3': values['ips3'], 'ips4': values['ips4'],
            'ipe1': values['ipe1'], 'ipe2': values['ipe2'], 'ipe3': values['ipe3'], 'ipe4': values['ipe4'],
            'save-xlsx': values['save-xlsx'],}

def get_my_ip():
    return socket.gethostbyname(socket.gethostname()).split('.')

def display_table(table, header):
    fn = tempfile.TemporaryFile().name
    with open(fn, 'w', newline='\n') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=header)
        writer.writeheader()
        for data in table:
            writer.writerow(data)
    ps = subprocess.Popen(["powershell", "-Command", f"Import-Csv -path {fn} | Out-GridView -Wait"],
                          creationflags=subprocess.CREATE_NO_WINDOW)

def clear_data(data: dict, header: list) -> dict:
    d = dict()
    for key in header:
        try:
            d[key] = data[key]
        except KeyError:
            d[key] = 'n/a'
    return d

def check_computer(ip: ipaddress, no: int, cfg: dict, table: list) -> None:
    global cnt_done
    if Func.detect_on(ip.__str__()):
        data = {'No': f'{no:03}', 'IP': ip.__str__()}
        for f in cfg['function_list']:
            data.update(f(ip.__str__()))
        data.update({'Time': datetime.datetime.now().isoformat(sep=' ')[:19]})
        lock.acquire()
        table.append(clear_data(data, cfg['property_list']))
        lock.release()
    lock.acquire()
    cnt_done = cnt_done + 1
    lock.release()

def save_xls(table, header):
    filename = sg.popup_get_file('Save as',
    title = None,
    default_path = "",
    default_extension = ".xlsx",
    save_as = True,
    multiple_files = False,
    file_types = (('Excel files', '*.xlsx'), ('ALL Files', '*.*')),
    no_window = True,
    no_titlebar = False,
    grab_anywhere = False,
    keep_on_top = True,
    location = (None, None),
    initial_folder = None,
    image = None,
    files_delimiter = ";",
    modal = True,
    history = False,
    history_setting_filename = None)
    if filename:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(header)
        for data in table:
            row=[]
            for field in header:
                row.append(data[field])
            ws.append(row)
        wb.save(filename)

def update_xlsx(table, header):
    wb = openpyxl.load_workbook(filename=config['DBfile'])
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == 'DB':
            break
    wb.active = s
    ws = wb.active
    for data in table:
        row = []
        for field in header:
            row.append(str(data[field]))
        exists = False
        for r in range(2, ws.max_row+1):
            if ws.cell(row=r, column=12).value == data['SerialNumber']: #column = column z s/n
                for c,d in enumerate(row, 1):
                    ws.cell(row=r, column=c).value = d
                exists = True
                break
        if not exists:
            ws.append(row)
    wb.save(config['DBfile'])


def database_update(table, header):
    if not ('Name' in header and 'MAC Address' and 'Last Logged User' in header):
        return
    con = sqlite3.connect('computers.sqlite')
    cur = con.cursor()
    try:
        cur.execute('''CREATE TABLE computers
                       (name text, ip text, mac text, user text, time text)''')
        con.commit()
    except:
        pass
    for record in table:
        cur.execute("INSERT INTO computers VALUES (" +
                    f"\'{record['Name']}\', \'{record['IP']}\', \'{record['MAC Address']}\', \'{record['Last Logged User']}\', \'{record['Time']}\')" )
    con.commit()
    con.close()

def get_config():
    global config
    try:
        with open(r'config.yaml', encoding='utf8') as yf:
            config = yaml.full_load(yf)
        config['ips'] = []
        for addr_range in config['IPranges']:
            ip_start, ip_end = addr_range.split('-')
            ip = ipaddress.ip_address(ip_start)
            while ip <= ipaddress.ip_address(ip_end):
                config['ips'].append(ip)
                ip = ip + 1
    except FileNotFoundError:
        print('brak pliku config.yaml')
        sys.exit(1)

cnt_done = 0

if __name__ == '__main__':
    cfg = get_params()
    get_config()
    if cfg.get('Update database', False):
        cfg['property_list'] = []
        for p in properties:
            cfg['property_list'].append(p.name)
        cfg['function_list'] = {p.func for p in properties if p.name in cfg['property_list']}
        cfg['property_list'] = ['No', 'IP'] + cfg['property_list'] + ['Time']
    else:
        if(cfg['use-iprange']):
            try:
                config['ips'] = []
                ip_start = ipaddress.ip_address(cfg['ips1'] + '.' + cfg['ips2'] + '.' + cfg['ips3'] + '.' + cfg['ips4'])
                ip_end = ipaddress.ip_address(cfg['ipe1'] + '.'  + cfg['ipe2'] + '.'  + cfg['ipe3'] + '.' +cfg['ipe4'])
                if ip_start > ip_end:
                    raise ValueError
                ip=ip_start
                while ip <= ipaddress.ip_address(ip_end):
                    config['ips'].append(ip)
                    ip = ip + 1
            except ValueError:
                sg.popup_ok('Błędny adres IP.', auto_close=True, auto_close_duration=3)
                sys.exit(1)
    table = []
    cnt_done = 0
    lock = threading.Lock()
    no = 1
    threads = []
    for ip in config['ips']:
        t = threading.Thread(target=check_computer, args=(ip, no, cfg, table,))
        threads.append(t)
        t.start()
        no = no + 1
        ip = ip + 1
    while threading.active_count() > 1:
        time.sleep(0.5)
        window['progress'].UpdateBar(cnt_done, len(config['ips']))
        print(threading.active_count())
    window['progress'].UpdateBar(1,1)
    time.sleep(0.3)
    for t in threads:
        t.join()
    window.close()
    table = sorted(table, key=lambda i:i['No'])
    for r in range(len(table)):
        table[r]['No'] = f'{r+1:03}'
    if cfg.get('save-xlsx', False):
        save_xls(table, cfg['property_list'])
    if cfg.get('Update database', False):
        update_xlsx(table, cfg['property_list'])
    display_table(table, cfg['property_list'])
